#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import time
import json
import csv
import random
import urllib.parse
import subprocess
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Set
from datetime import datetime
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

# Excel equivalence ATC (optionnel si tu veux remplir "Libellé ATC")
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None  # noqa

# ============================================================
# CONFIG
# ============================================================

# BDPM (TXT)
BDPM_CIS_URL = "https://base-donnees-publique.medicaments.gouv.fr/download/file/CIS_bdpm.txt"
BDPM_CIS_CIP_URL = "https://base-donnees-publique.medicaments.gouv.fr/download/file/CIS_CIP_bdpm.txt"
BDPM_COMPO_URL = "https://base-donnees-publique.medicaments.gouv.fr/download/file/CIS_COMPO_bdpm.txt"

# BDPM "Médicaments d'intérêt thérapeutique majeur" (CIS -> ATC)
BDPM_MITM_URL = "https://base-donnees-publique.medicaments.gouv.fr/download/file/CIS_MITM.txt"

# BDPM "Informations importantes" (génération en direct) (CIS -> URL info importante)
BDPM_INFO_IMPORTANTES_URL = "https://base-donnees-publique.medicaments.gouv.fr/download/CIS_InfoImportantes.txt"

# Page BDPM pour fiche-info (HTML)
BDPM_DOC_EXTRACT_URL = "https://base-donnees-publique.medicaments.gouv.fr/medicament/{cis}/extrait"

# ANSM rétrocession
ANSM_RETRO_PAGE = "https://ansm.sante.fr/documents/reference/medicaments-en-retrocession"

# Airtable
AIRTABLE_API_BASE = "https://api.airtable.com/v0"
AIRTABLE_MIN_DELAY_S = float(os.getenv("AIRTABLE_MIN_DELAY_S", "0.25"))
AIRTABLE_BATCH_SIZE = 10
UPDATE_FLUSH_THRESHOLD = int(os.getenv("UPDATE_FLUSH_THRESHOLD", "200"))

HTTP_CONNECT_TIMEOUT = float(os.getenv("HTTP_CONNECT_TIMEOUT", "10"))
HTTP_READ_TIMEOUT = float(os.getenv("HTTP_READ_TIMEOUT", "25"))
REQUEST_TIMEOUT = 35
MAX_RETRIES = 4

REPORT_DIR = os.getenv("REPORT_DIR", "reports")
REPORT_COMMIT = os.getenv("GITHUB_COMMIT_REPORT", "0").strip() == "1"
HEARTBEAT_EVERY = int(os.getenv("HEARTBEAT_EVERY", "50"))

# Fichier Excel d'équivalence ATC (niveau 4 -> libellé)
ATC_EQUIVALENCE_FILE = os.getenv("ATC_EQUIVALENCE_FILE", "data/equivalence atc.xlsx")

HEADERS_WEB = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123 Safari/537.36",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.7",
}

# ============================================================
# AIRTABLE FIELDS (adapte ici si besoin)
# ============================================================

FIELD_CIS = "Code cis"
FIELD_RCP = "Lien vers RCP"
FIELD_CIP13 = "CIP 13"
FIELD_DISPO = "Disponibilité du traitement"
FIELD_CPD = "Conditions de prescription et délivrance"
FIELD_LABO = "Laboratoire"
FIELD_SPEC = "Spécialité"
FIELD_FORME = "Forme"
FIELD_VOIE = "Voie d'administration"
FIELD_ATC = "Code ATC"
FIELD_ATC4 = "Code ATC (niveau 4)"      # computed -> lecture seulement
FIELD_ATC_LABEL = "Libellé ATC"          # champ à écrire
FIELD_COMPOSITION = "Composition"        # champ à écrire
FIELD_COMPOSITION_DETAILS = "Composition détails"  # champ à écrire (sans ligne "sans accents")
FIELD_LIEN_INFO_IMPORTANTE = "Lien vers information importante"  # champ à écrire (URL)

# ✅ Moment de prise : on veut UNIQUEMENT les mots-clés présents (pas la phrase)
FIELD_MOMENT_PRISE = "Moment de prise"  # champ à écrire

# ✅ Champs RCP à extraire dans Airtable
FIELD_INDICATIONS_RCP = "Indications RCP"      # 4.1
FIELD_POSOLOGIE_RCP = "Posologie RCP"          # 4.2
FIELD_INTERACTIONS_RCP = "Interactions RCP"    # 4.5

# ✅ timestamp de revue
FIELD_DATE_REVUE = "Date revue ligne"   # champ à écrire

# règle absolue : ne jamais écrire ces champs (computed)
DO_NOT_WRITE_FIELDS: Set[str] = {
    FIELD_ATC4,
}

# ============================================================
# LOG
# ============================================================

def _ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")

def die(msg: str, code: int = 1):
    print(f"[{_ts()}] ❌ {msg}", flush=True)
    raise SystemExit(code)

def info(msg: str):
    print(f"[{_ts()}] ℹ️ {msg}", flush=True)

def ok(msg: str):
    print(f"[{_ts()}] ✅ {msg}", flush=True)

def warn(msg: str):
    print(f"[{_ts()}] ⚠️ {msg}", flush=True)

def sleep_throttle():
    time.sleep(AIRTABLE_MIN_DELAY_S)

def retry_sleep(attempt: int):
    time.sleep(min(10, 0.6 * (2 ** (attempt - 1))) + random.random() * 0.25)

# ============================================================
# REVIEW TIMESTAMP
# ============================================================

def now_paris_iso_seconds() -> str:
    return datetime.now(ZoneInfo("Europe/Paris")).isoformat(timespec="seconds")

# ============================================================
# REPORTING
# ============================================================

def ensure_report_dir():
    os.makedirs(REPORT_DIR, exist_ok=True)

def run_id_paris() -> str:
    return datetime.now(ZoneInfo("Europe/Paris")).strftime("%Y%m%d_%H%M%S")

def report_path_deleted_today() -> str:
    ensure_report_dir()
    fname = f"deleted_records_{time.strftime('%Y-%m-%d')}.txt"
    return os.path.join(REPORT_DIR, fname)

def append_deleted_report(cis: str, reason: str, url: str):
    p = report_path_deleted_today()
    line = f"{_ts()}\tCIS={cis}\tSUPPRIME\treason={reason}\turl={url}\n"
    with open(p, "a", encoding="utf-8") as f:
        f.write(line)

def init_moment_report_csv(path: str):
    ensure_report_dir()
    if not os.path.exists(path):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["run_id", "cis", "rcp_url", "status", "keywords"])

def append_moment_report(path: str, run_id: str, cis: str, rcp_url: str, status: str, keywords: str):
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([run_id, cis, rcp_url, status, keywords])
        f.flush()

def try_git_commit_reports():
    """
    Commit/push optionnel des rapports dans REPORT_DIR (si REPORT_COMMIT=1).
    """
    if not REPORT_COMMIT:
        return
    try:
        if not os.path.exists(REPORT_DIR):
            return
        subprocess.run(["git", "status"], check=False)
        subprocess.run(["git", "add", REPORT_DIR], check=True)
        subprocess.run(["git", "commit", "-m", f"Reports: {time.strftime('%Y-%m-%d %H:%M:%S')}"], check=True)
        subprocess.run(["git", "push"], check=True)
        ok("Rapports commit/push sur GitHub effectués.")
    except Exception as e:
        warn(f"Commit/push des rapports impossible (on continue): {e}")

# ============================================================
# TEXT UTIL
# ============================================================

def safe_text(s: str) -> str:
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = s.replace("\uFFFD", "")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return s.strip()

def strip_accents(s: str) -> str:
    s = safe_text(s)
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )

def normalize_ws_keep_lines(s: str) -> str:
    s = safe_text(s)
    lines = []
    for line in s.split("\n"):
        line = re.sub(r"[ \t]{2,}", " ", line).strip()
        lines.append(line)
    out = []
    empty = 0
    for line in lines:
        if line == "":
            empty += 1
            if empty <= 1:
                out.append("")
        else:
            empty = 0
            out.append(line)
    return "\n".join(out).strip()

def capitalize_each_line(text: str) -> str:
    if not text:
        return text
    out_lines: List[str] = []
    for line in text.split("\n"):
        stripped = line.lstrip()
        if stripped:
            prefix_len = len(line) - len(stripped)
            prefix = line[:prefix_len]
            stripped = stripped[0].upper() + stripped[1:]
            out_lines.append(prefix + stripped)
        else:
            out_lines.append(line)
    return "\n".join(out_lines)

def chunked(lst: List, n: int):
    for i in range(0, len(lst), n):
        yield lst[i:i+n]

def _bs_parser():
    try:
        import lxml  # noqa: F401
        return "lxml"
    except Exception:
        return "html.parser"

# ============================================================
# EXTRACTION "MOMENT DE PRISE" (RCP)
# ============================================================

# ✅ Mots-clés EXACTS à récupérer si présents dans le RCP
# (on ne stocke PAS la phrase, uniquement les mots-clés trouvés)
MOMENT_PRISE_KEYWORDS = [
    # Formulations fréquentes des RCP
    "avec un repas",
    "avec les repas",
    "au cours du repas",
    "pendant le repas",
    "avant le repas",
    "apres le repas",
    "après le repas",
    "en dehors des repas",
    "au moment des repas",

    # Avec / sans nourriture
    "avec nourriture",
    "avec de la nourriture",
    "avec des aliments",
    "prendre avec nourriture",
    "avec ou sans nourriture",
    "avec ou sans aliments",
    "sans nourriture",

    # Interactions liées à la nourriture
    "repas riche",
    "repas riche en graisses",
    "repas riche en graisse",
    "repas gras",
]

def _norm(s: str) -> str:
    return strip_accents(safe_text(s)).lower()

# Trie par longueur décroissante : si plusieurs expressions coexistent, on privilégie les plus spécifiques
_KW_NORM_SORTED = sorted([_norm(k) for k in MOMENT_PRISE_KEYWORDS if _norm(k)], key=len, reverse=True)

def _build_kw_patterns(keywords_norm: List[str]) -> List[Tuple[str, re.Pattern]]:
    """
    Regex robustes sur texte normalisé (sans accents):
    - espaces -> \\s+ (gère retours ligne et espaces multiples)
    - bornes de mots \\b pour limiter les faux positifs
    """
    out: List[Tuple[str, re.Pattern]] = []
    for kw in keywords_norm:
        kw_esc = re.escape(kw)
        kw_esc = re.sub(r"\\\s+", r"\\s+", kw_esc)
        rgx = re.compile(rf"\b{kw_esc}\b", flags=re.IGNORECASE)
        out.append((kw, rgx))
    return out

_KW_PATTERNS = _build_kw_patterns(_KW_NORM_SORTED)

def extract_moment_keywords_from_rcp_html(html: str) -> str:
    """
    Retourne uniquement les mots-clés présents (distincts), séparés par '; '.
    Si aucun mot-clé: retourne "" (et on n'écrit rien dans Airtable).
    """
    if not html:
        return ""

    soup = BeautifulSoup(html, _bs_parser())
    raw = soup.get_text(" ", strip=True)
    if not raw:
        return ""

    t = _norm(raw)

    found: List[str] = []
    for kw_norm, rgx in _KW_PATTERNS:
        if rgx.search(t):
            found.append(kw_norm)

    # Dé-doublonnage, conservation de l'ordre (déjà trié par spécificité)
    seen = set()
    uniq_norm: List[str] = []
    for f in found:
        if f in seen:
            continue
        uniq_norm.append(f)
        seen.add(f)

    if not uniq_norm:
        return ""

    # On renvoie les libellés originaux (ceux de MOMENT_PRISE_KEYWORDS) si possible.
    # On mappe norm -> original (première occurrence).
    norm_to_original: Dict[str, str] = {}
    for orig in MOMENT_PRISE_KEYWORDS:
        n = _norm(orig)
        if n and n not in norm_to_original:
            norm_to_original[n] = orig

    rendered = [norm_to_original.get(n, n) for n in uniq_norm]
    # sécurité anti-champs énormes
    if len(rendered) > 30:
        rendered = rendered[:30]
    return "; ".join(rendered).strip()


# ============================================================
# EXTRACTION RCP SECTIONS 4.1 / 4.2 / 4.5
# ============================================================

def _normalize_rcp_lines(html: str) -> List[str]:
    soup = BeautifulSoup(html, _bs_parser())
    raw = soup.get_text("\n", strip=True)
    raw = safe_text(raw)
    if not raw:
        return []
    lines: List[str] = []
    for ln in raw.split("\n"):
        ln = re.sub(r"\s+", " ", ln).strip()
        if ln:
            lines.append(ln)
    return lines

def _find_section_start(lines: List[str], sec: str) -> Optional[int]:
    pat = re.compile(rf"^\s*{re.escape(sec)}\b")
    for i, ln in enumerate(lines):
        if pat.match(ln):
            return i
    return None

def _find_section_end(lines: List[str], start_idx: int, stop_pat: re.Pattern) -> int:
    for j in range(start_idx + 1, len(lines)):
        if stop_pat.match(lines[j]):
            return j
    return len(lines)

def _extract_section(lines: List[str], sec: str, stop_pat: re.Pattern) -> str:
    start = _find_section_start(lines, sec)
    if start is None:
        return ""
    end = _find_section_end(lines, start, stop_pat)

    body_lines = lines[start + 1:end]

    title_ln = lines[start]
    title_tail = re.sub(rf"^\s*{re.escape(sec)}\s*", "", title_ln).strip()
    if ":" in title_tail:
        after = title_tail.split(":", 1)[1].strip()
        if after:
            body_lines = [after] + body_lines

    text = "\n".join(body_lines).strip()
    return normalize_ws_keep_lines(text)

def extract_rcp_sections_from_html(html: str) -> Dict[str, str]:
    if not html:
        return {"4.1": "", "4.2": "", "4.5": ""}

    lines = _normalize_rcp_lines(html)
    if not lines:
        return {"4.1": "", "4.2": "", "4.5": ""}

    stop_41 = re.compile(r"^\s*(?:4\.2\b|5\.)")
    stop_42 = re.compile(r"^\s*(?:4\.[3-9]\b|5\.)")
    stop_45 = re.compile(r"^\s*(?:4\.[6-9]\b|5\.)")

    return {
        "4.1": _extract_section(lines, "4.1", stop_41),
        "4.2": _extract_section(lines, "4.2", stop_42),
        "4.5": _extract_section(lines, "4.5", stop_45),
    }

# ============================================================
# ATC HELPERS (pour lecture fichiers)
# ============================================================

ATC7_PAT = re.compile(r"^[A-Z]\d{2}[A-Z]{2}\d{2}$")  # ex A11CA01
ATC5_PAT = re.compile(r"^[A-Z]\d{2}[A-Z]{2}$")       # ex A11CA

def canonical_atc7(raw: str) -> str:
    if not raw:
        return ""
    s = re.sub(r"[^A-Za-z0-9]", "", raw).upper()
    return s if ATC7_PAT.fullmatch(s) else ""

def atc_level4_from_any(atc: str) -> str:
    a = (atc or "").strip().upper().replace(" ", "")
    if ATC7_PAT.fullmatch(a):
        return a[:5]
    if ATC5_PAT.fullmatch(a):
        return a
    m7 = re.search(r"\b([A-Z]\d{2}[A-Z]{2}\d{2})\b", a)
    if m7:
        return m7.group(1)[:5]
    m5 = re.search(r"\b([A-Z]\d{2}[A-Z]{2})\b", a)
    if m5:
        return m5.group(1)
    return ""

def load_atc_equivalence_excel(path: str) -> Dict[str, str]:
    if not os.path.exists(path):
        warn(f"Fichier d'équivalence ATC introuvable: {path} (Libellé ATC restera vide)")
        return {}
    if pd is None:
        warn("pandas non disponible -> impossible de lire l'Excel d'équivalence ATC (Libellé ATC restera vide)")
        return {}
    try:
        df = pd.read_excel(path)
    except Exception as e:
        warn(f"Impossible de lire l'Excel {path}: {e} (Libellé ATC restera vide)")
        return {}

    if FIELD_ATC4 not in df.columns or FIELD_ATC_LABEL not in df.columns:
        warn(
            f"Colonnes attendues absentes dans {path}. "
            f"Trouvé: {list(df.columns)} | Attendu: ['{FIELD_ATC4}', '{FIELD_ATC_LABEL}']"
        )
        return {}

    mapping: Dict[str, str] = {}
    for _, row in df.iterrows():
        code = safe_text(row.get(FIELD_ATC4, "")).upper()
        label = safe_text(row.get(FIELD_ATC_LABEL, ""))
        if not code or not label:
            continue
        code = atc_level4_from_any(code) or code
        mapping[code] = label

    ok(f"Équivalence ATC chargée: {len(mapping)} entrées")
    return mapping

# ============================================================
# DOWNLOAD
# ============================================================

def http_get(url: str, timeout: Tuple[float, float] = (HTTP_CONNECT_TIMEOUT, 60.0)) -> requests.Response:
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, headers=HEADERS_WEB, timeout=timeout, allow_redirects=True)
            return r
        except Exception as e:
            last_err = e
            retry_sleep(attempt)
    raise RuntimeError(f"GET failed: {url} / {last_err}")

def download_text(url: str, encoding: str = "latin-1") -> str:
    r = http_get(url, timeout=(HTTP_CONNECT_TIMEOUT, 120.0))
    if r.status_code >= 400:
        raise RuntimeError(f"HTTP {r.status_code} for {url}")
    r.encoding = encoding
    return r.text

def download_bytes(url: str, timeout_s: float = 140.0) -> bytes:
    r = http_get(url, timeout=(HTTP_CONNECT_TIMEOUT, timeout_s))
    if r.status_code >= 400:
        raise RuntimeError(f"HTTP {r.status_code} for {url}")
    return r.content

# ============================================================
# ANSM retrocession
# ============================================================

def find_ansm_retro_excel_link() -> str:
    r = http_get(ANSM_RETRO_PAGE, timeout=(HTTP_CONNECT_TIMEOUT, 60.0))
    if r.status_code >= 400:
        raise RuntimeError(f"HTTP {r.status_code} {ANSM_RETRO_PAGE}")
    soup = BeautifulSoup(r.text, _bs_parser())

    links = []
    for a in soup.find_all("a", href=True):
        href = (a["href"] or "").strip()
        if not href:
            continue
        if href.startswith("/"):
            href = "https://ansm.sante.fr" + href
        low = href.lower()
        if ("ansm.sante.fr/uploads/" in low) and ("retrocession" in low) and re.search(r"\.xlsx?$|\.xls$", low):
            links.append(href)

    if not links:
        raise RuntimeError("Lien Excel ANSM (rétrocession) introuvable sur la page")

    def score(u: str) -> Tuple[int, str]:
        m = re.search(r"/(\d{4})/(\d{2})/(\d{2})/", u)
        if m:
            return (1, f"{m.group(1)}{m.group(2)}{m.group(3)}")
        return (0, u)

    links.sort(key=score, reverse=True)
    return links[0]

def parse_ansm_retrocession_cis(excel_bytes: bytes, url_hint: str = "") -> Set[str]:
    cis_set: Set[str] = set()
    ext = ""
    if url_hint:
        ext = url_hint.lower().split("?")[0].split("#")[0]
        ext = os.path.splitext(ext)[1].lower()

    import io
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue
            v = row[2]
            if v is None:
                continue
            v = re.sub(r"\D", "", str(v))
            if len(v) == 8:
                cis_set.add(v)
        return cis_set

    try:
        import xlrd  # type: ignore
    except Exception:
        raise RuntimeError("Le fichier ANSM est en .xls mais 'xlrd' n'est pas installé. pip install xlrd==2.0.1")

    book = xlrd.open_workbook(file_contents=excel_bytes)
    sheet = book.sheet_by_index(0)
    for rx in range(sheet.nrows):
        row = sheet.row_values(rx)
        if len(row) < 3:
            continue
        v = row[2]
        if v is None:
            continue
        v = re.sub(r"\D", "", str(v))
        if len(v) == 8:
            cis_set.add(v)

    return cis_set

# ============================================================
# BDPM PARSE (CIS, CIP)
# ============================================================

@dataclass
class CisRow:
    cis: str
    specialite: str
    forme: str
    voie_admin: str
    titulaire: str

def parse_bdpm_cis(txt: str) -> Dict[str, CisRow]:
    out: Dict[str, CisRow] = {}
    for line in txt.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 4:
            continue
        cis = re.sub(r"\D", "", parts[0].strip())
        if len(cis) != 8:
            continue
        denom = safe_text(parts[1]) if len(parts) > 1 else ""
        forme = safe_text(parts[2]) if len(parts) > 2 else ""
        voie = safe_text(parts[3]) if len(parts) > 3 else ""
        titulaire = safe_text(parts[10]) if len(parts) > 10 else ""
        out[cis] = CisRow(cis=cis, specialite=denom, forme=forme, voie_admin=voie, titulaire=titulaire)
    return out

@dataclass
class CipInfo:
    cip13: str
    has_taux: bool

def looks_like_taux(val: str) -> bool:
    v = (val or "").strip()
    if not v:
        return False
    v2 = v.replace(",", ".").replace("%", "").strip()
    if not re.fullmatch(r"\d{1,3}(\.\d+)?", v2):
        return False
    try:
        x = float(v2)
    except Exception:
        return False
    return x in {0, 15, 30, 35, 65, 100}

def parse_bdpm_cis_cip(txt: str) -> Dict[str, CipInfo]:
    out: Dict[str, CipInfo] = {}
    for line in txt.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        cis = re.sub(r"\D", "", parts[0].strip())
        if len(cis) != 8:
            continue

        cip13 = ""
        for p in parts:
            d = re.sub(r"\D", "", p)
            if len(d) == 13:
                cip13 = d
                break

        has_taux = any(looks_like_taux(p) for p in parts)

        if cis not in out:
            out[cis] = CipInfo(cip13=cip13, has_taux=has_taux)
        else:
            if not out[cis].cip13 and cip13:
                out[cis].cip13 = cip13
            out[cis].has_taux = out[cis].has_taux or has_taux

    return out

def normalize_lab_name(titulaire: str) -> str:
    t = titulaire or ""
    t = t.replace(",", " ").replace(";", " ")
    t = re.sub(r"\s+", " ", t).strip()
    if not t:
        return ""
    t = re.sub(r"\b(SAS|SA|SARL|S\.A\.|S\.A\.S\.|GMBH|LTD|INC|BV|AG|SPA|S\.P\.A\.)\b", "", t, flags=re.IGNORECASE).strip()
    t = re.sub(r"\s+", " ", t).strip()
    first = t.split(" ")[0].strip()
    if first.isupper() and len(first) > 2:
        first = first.capitalize()
    if first.lower() in {"laboratoires", "laboratoire"} and len(t.split(" ")) > 1:
        nxt = t.split(" ")[1]
        if nxt.isupper() and len(nxt) > 2:
            nxt = nxt.capitalize()
        return nxt
    return first

# ============================================================
# COMPOSITION BDPM -> DCI principales
# ============================================================

_NOISE_RE = re.compile(r"\b(pour\s+pr[ée]parations?\s+hom[ée]opathiques)\b", flags=re.IGNORECASE)
_COMPLEX_PREFIX_RE = re.compile(r"^\s*complexe\s+d['’]\s*", flags=re.IGNORECASE)

_SALT_WORDS = (
    r"chlorhydrate|dichlorhydrate|bromhydrate|chlorure|bromure|iodure|fluorure|"
    r"citrate|fumarate|succinate|tartrate|mal[eé]ate|m[eé]silate|mesilate|"
    r"tosylate|benzoate|gluconate|lactate|carbonate|phosphate|sulfate|sulphate|"
    r"ac[eé]tate|oxalate|nitrate|nitrite|hydrog[eé]nosuccinate|"
    r"b[eé]silate|besilate|besylate"
)

_SALT_PREFIX_RE = re.compile(rf"\b({_SALT_WORDS})\s+(?:de|d['’])\s*", flags=re.IGNORECASE)
_SALT_LEADING_RE = re.compile(rf"^\s*({_SALT_WORDS})\s+", flags=re.IGNORECASE)
_SALT_GLUE_RE = re.compile(rf"^({_SALT_WORDS})([a-zà-ÿ])", flags=re.IGNORECASE)

_PREFIX_GLUE_RES = [
    re.compile(r"^(dichlorhydrate)([a-zà-ÿ])", re.IGNORECASE),
    re.compile(r"^(chlorhydrate)([a-zà-ÿ])", re.IGNORECASE),
    re.compile(r"^(dioxyd(e|e)\s*|dioxyde)([a-zà-ÿ])", re.IGNORECASE),
    re.compile(r"^(dioxyde)([a-zà-ÿ])", re.IGNORECASE),
    re.compile(r"^(peroxyde)([a-zà-ÿ])", re.IGNORECASE),
    re.compile(r"^(oxyde)([a-zà-ÿ])", re.IGNORECASE),
]

_HYDRATE_RE = re.compile(
    r"\b("
    r"anhydre|base|"
    r"sesquihydrat[ée]?|"
    r"monohydrat[ée]?|dihydrat[ée]?|trihydrat[ée]?|t[ée]trahydrat[ée]?|"
    r"pentahydrat[ée]?|h[ée]mihydrat[ée]?|hydrat[ée]?"
    r")\b",
    flags=re.IGNORECASE
)

_COUNTERION_TAIL_RE = re.compile(
    r"\b("
    r"arginine|sodique|sodium|potassique|potassium|calcique|calcium|magnesium|magn[eé]sium|lithium|ammonium|"
    r"zinc|cuivre|aluminium|manganese|manganèse"
    r")\b",
    flags=re.IGNORECASE
)

_DESC_TAIL_RE = re.compile(
    r"\b("
    r"humain|humaine|biog[eé]n[eé]tique|recombinant|recombinante|"
    r"biosynth[eé]tique|biotechnologique|analogue|synthetique|synth[eé]tique"
    r")\b",
    flags=re.IGNORECASE
)

def _pretty_segment(s: str) -> str:
    s = safe_text(s)
    if not s:
        return ""
    s = s.lower()
    return s[0].upper() + s[1:] if s else ""

def clean_to_main_dci(raw: str) -> str:
    s = safe_text(raw)
    if not s:
        return ""

    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"\([^)]*\)", " ", s)

    s = s.replace("\\", " ")
    s = s.replace("/", " / ")
    s = re.sub(r"\s+", " ", s).strip()

    s = _COMPLEX_PREFIX_RE.sub("", s)
    s = _NOISE_RE.sub(" ", s)

    for rgx in _PREFIX_GLUE_RES:
        s = rgx.sub(r"\1 \2", s)

    s = _SALT_GLUE_RE.sub(r"\1 \2", s)

    if re.match(r"^\s*(dioxyde|oxyde|peroxyde)\b", s, flags=re.IGNORECASE):
        return ""

    s = _SALT_PREFIX_RE.sub("", s)
    s = _SALT_LEADING_RE.sub("", s)
    s = _HYDRATE_RE.sub(" ", s)
    s = _DESC_TAIL_RE.sub(" ", s)

    s = re.sub(r"[;,:]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    tokens = s.split()
    while tokens and _COUNTERION_TAIL_RE.fullmatch(tokens[-1]):
        tokens.pop()
    s = " ".join(tokens).strip()
    if not s:
        return ""

    return _pretty_segment(s)

def parse_bdpm_compositions(txt: str) -> Dict[str, str]:
    cis_to_set: Dict[str, Dict[str, str]] = {}

    for line in txt.splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 4:
            continue

        cis = re.sub(r"\D", "", parts[0].strip())
        if len(cis) != 8:
            continue

        denom = safe_text(parts[3])
        if not denom:
            continue

        denom_norm = denom.replace("\\|", "|").replace("|", "|")
        denom_norm = denom_norm.replace("/", "|")
        pieces = [p.strip() for p in denom_norm.split("|") if p.strip()]

        for piece in pieces:
            dci = clean_to_main_dci(piece)
            if not dci:
                continue
            key = dci.lower().strip()
            cis_to_set.setdefault(cis, {})
            cis_to_set[cis][key] = dci

    out: Dict[str, str] = {}
    for cis, kv in cis_to_set.items():
        values = list(kv.values())
        values.sort(key=lambda x: x.lower())
        out[cis] = " - ".join(values)

    ok(f"Compositions (DCI principales) chargées: {len(out)} CIS")
    return out

# ============================================================
# BDPM MITM (CIS -> ATC)
# ============================================================

def parse_mitm_cis_to_atc(txt: str) -> Dict[str, str]:
    cis_to_atc: Dict[str, str] = {}
    for line in (txt or "").splitlines():
        if not line.strip():
            continue
        cis_m = re.search(r"\b(\d{8})\b", line)
        if not cis_m:
            continue
        cis = cis_m.group(1)
        atc_m = re.search(r"\b([A-Z]\d{2}[A-Z]{2}\d{2})\b", line.upper())
        if not atc_m:
            continue
        atc = canonical_atc7(atc_m.group(1))
        if atc:
            cis_to_atc[cis] = atc
    ok(f"MITM (ATC) chargé: {len(cis_to_atc)} correspondances CIS->ATC")
    return cis_to_atc

# ============================================================
# BDPM Informations importantes (CIS -> URL)
# ============================================================

_URL_PAT = re.compile(r"(https?://[^\s\"'<>]+)", re.IGNORECASE)

def parse_info_importantes_cis_to_url(txt: str) -> Dict[str, str]:
    cis_to_url: Dict[str, str] = {}
    for line in (txt or "").splitlines():
        if not line.strip():
            continue
        cis_m = re.search(r"\b(\d{8})\b", line)
        if not cis_m:
            continue
        cis = cis_m.group(1)
        um = _URL_PAT.search(line)
        if not um:
            continue
        url = um.group(1).strip().rstrip(").,;")
        if url.startswith("http"):
            cis_to_url[cis] = url
    ok(f"Infos importantes chargées: {len(cis_to_url)} correspondances CIS->URL")
    return cis_to_url

# ============================================================
# URL HELPERS
# ============================================================

def base_extrait_url_from_cis(cis: str) -> str:
    return BDPM_DOC_EXTRACT_URL.format(cis=cis)

def set_tab(url: str, cis_fallback: str, tab: str) -> str:
    if not url or not url.startswith("http"):
        url = base_extrait_url_from_cis(cis_fallback)
    parts = urllib.parse.urlsplit(url)
    qs = urllib.parse.parse_qs(parts.query, keep_blank_values=True)
    qs["tab"] = [tab]
    new_query = urllib.parse.urlencode(qs, doseq=True)

    frag = parts.fragment or ""
    if re.search(r"\btab=", frag):
        frag = re.sub(r"tab=[^&]+", f"tab={tab}", frag)
    else:
        frag = f"tab={tab}"

    cleaned = parts._replace(query=new_query, fragment=frag)
    return urllib.parse.urlunsplit(cleaned)

def rcp_link_default(cis: str) -> str:
    return f"{base_extrait_url_from_cis(cis)}#tab=rcp"

# ============================================================
# FICHE-INFO SCRAPING
# ============================================================

HOMEOPATHY_PAT = re.compile(r"hom[ée]opath(?:ie|ique)", flags=re.IGNORECASE)
HOMEOPATHY_CLASS_PAT = re.compile(r"m[ée]dicament\s+hom[ée]opathique", flags=re.IGNORECASE)

RESERVED_HOSP_PAT = re.compile(r"réserv[ée]?\s+à\s+l['’]usage\s+hospitalier", flags=re.IGNORECASE)
USAGE_HOSP_PAT = re.compile(r"\busage\s+hospitalier\b", flags=re.IGNORECASE)
NEGATION_PAT = re.compile(
    r"(?:\bnon\b|\bpas\b|\bjamais\b)\s+(?:réserv[ée]?\s+à\s+l['’]usage\s+hospitalier|\busage\s+hospitalier\b)",
    flags=re.IGNORECASE
)

GLOSSARY_PAT = re.compile(r"\baller\s+au\s+glossaire\b", flags=re.IGNORECASE)

class PageUnavailable(Exception):
    def __init__(self, url: str, status: Optional[int], detail: str):
        super().__init__(detail)
        self.url = url
        self.status = status
        self.detail = detail

def clean_cpd_text_keep_useful(text: str) -> str:
    if not text:
        return ""
    lines = [ln.strip() for ln in text.split("\n")]
    kept = []
    for ln in lines:
        if not ln:
            kept.append("")
            continue
        if GLOSSARY_PAT.search(ln):
            continue
        kept.append(ln)
    return normalize_ws_keep_lines("\n".join(kept))

def fetch_html_checked(url: str, timeout: Tuple[float, float] = (HTTP_CONNECT_TIMEOUT, HTTP_READ_TIMEOUT), max_retries: int = 3) -> str:
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            r = requests.get(url, headers=HEADERS_WEB, timeout=timeout, allow_redirects=True)
            if r.status_code == 404:
                raise PageUnavailable(url, 404, "HTTP 404")
            if r.status_code >= 400:
                raise PageUnavailable(url, r.status_code, f"HTTP {r.status_code}")
            if not r.text or len(r.text) < 200:
                raise PageUnavailable(url, r.status_code, "HTML vide/trop court")
            return r.text
        except PageUnavailable:
            raise
        except Exception as e:
            last_err = e
            time.sleep(1.0 * attempt)
    raise PageUnavailable(url, None, f"Erreur réseau: {last_err}")

def detect_homeopathy_from_fiche_info(soup: BeautifulSoup) -> bool:
    text = soup.get_text("\n", strip=True)
    return bool(HOMEOPATHY_PAT.search(text) or HOMEOPATHY_CLASS_PAT.search(text))

def extract_badge_usage_hospitalier_only(soup: BeautifulSoup) -> bool:
    for el in soup.find_all(["span", "div", "a", "p", "li"]):
        t = (el.get_text(" ", strip=True) or "")
        if not t:
            continue
        if len(t) > 60:
            continue
        if "cela signifie" in t.lower():
            continue
        if t.strip().lower() == "usage hospitalier":
            return True
    return False

def extract_cpd_from_fiche_info(soup: BeautifulSoup) -> str:
    lines = [ln.strip() for ln in soup.get_text("\n", strip=True).split("\n")]

    autres_pat = re.compile(r"^Autres\s+informations$", re.IGNORECASE)
    cpd_pat = re.compile(r"^Conditions\s+de\s+prescription\s+et\s+de\s+d[ée]livrance\b", re.IGNORECASE)
    stop_pat = re.compile(
        r"^(Statut\s+de\s+l['’]autorisation|Type\s+de\s+proc[ée]dure|Code\s+CIS|Titulaire\s+de\s+l['’]autorisation)\s*:",
        re.IGNORECASE
    )

    start_autres = None
    for i, ln in enumerate(lines):
        if autres_pat.match(ln):
            start_autres = i
            break
    if start_autres is None:
        return ""

    start_cpd = None
    inline_value = ""
    for i in range(start_autres, len(lines)):
        ln = lines[i]
        if cpd_pat.match(ln):
            start_cpd = i
            if ":" in ln:
                inline_value = ln.split(":", 1)[1].strip()
            break
    if start_cpd is None:
        return ""

    collected: List[str] = []
    if inline_value:
        collected.append(inline_value)

    for ln in lines[start_cpd + 1:]:
        if stop_pat.search(ln):
            break
        collected.append(ln)

    return clean_cpd_text_keep_useful(normalize_ws_keep_lines("\n".join(collected)))

def analyze_fiche_info(fiche_url: str) -> Tuple[str, bool, bool, bool]:
    html = fetch_html_checked(fiche_url)
    soup = BeautifulSoup(html, _bs_parser())

    is_homeo = detect_homeopathy_from_fiche_info(soup)

    cpd_text = extract_cpd_from_fiche_info(soup)
    cpd_text = capitalize_each_line(cpd_text)

    badge_usage = extract_badge_usage_hospitalier_only(soup)

    zone_text = cpd_text or ""
    if NEGATION_PAT.search(zone_text):
        reserved = False
        usage = False
    else:
        reserved = bool(RESERVED_HOSP_PAT.search(zone_text))
        usage = bool(USAGE_HOSP_PAT.search(zone_text)) or badge_usage

    return cpd_text, is_homeo, reserved, usage

# ============================================================
# DISPONIBILITE
# ============================================================

def compute_disponibilite(
    has_taux_ville: bool,
    is_ansm_retro: bool,
    is_homeo: bool,
    reserved_hospital: bool,
    usage_hospital: bool
) -> str:
    ville = bool(has_taux_ville or is_homeo)

    if is_ansm_retro and ville:
        return "Disponible en pharmacie de ville et en rétrocession hospitalière"
    if is_ansm_retro and not ville:
        return "Disponible en rétrocession hospitalière"

    if reserved_hospital:
        return "Réservé à l'usage hospitalier"
    if usage_hospital and not ville:
        return "Réservé à l'usage hospitalier"

    if ville:
        return "Disponible en pharmacie de ville"

    return "Pas d'information sur la disponibilité mentionnée"

# ============================================================
# AIRTABLE CLIENT
# ============================================================

class AirtableClient:
    def __init__(self, api_token: str, base_id: str, table_name: str):
        self.api_token = api_token
        self.base_id = base_id
        self.table_name = table_name
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {self.api_token}",
            "Content-Type": "application/json",
        })

    @property
    def table_url(self) -> str:
        t = urllib.parse.quote(self.table_name, safe="")
        return f"{AIRTABLE_API_BASE}/{self.base_id}/{t}"

    def _request(self, method: str, url: str, **kwargs):
        last_err = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                sleep_throttle()
                r = self.session.request(method, url, timeout=REQUEST_TIMEOUT, **kwargs)
                if r.status_code in (429, 500, 502, 503, 504):
                    raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
                if r.status_code >= 400:
                    raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
                return r
            except Exception as e:
                last_err = e
                retry_sleep(attempt)
        raise RuntimeError(f"Airtable request failed: {method} {url} / {last_err}")

    def list_all_records(self, fields: Optional[List[str]] = None) -> List[dict]:
        requested_fields = list(fields) if fields else None

        while True:
            out: List[dict] = []
            params = {}
            if requested_fields:
                params["fields[]"] = requested_fields

            offset = None
            try:
                while True:
                    if offset:
                        params["offset"] = offset
                    r = self._request("GET", self.table_url, params=params)
                    data = r.json()
                    out.extend(data.get("records", []))
                    offset = data.get("offset")
                    if not offset:
                        break
                return out
            except Exception as e:
                msg = str(e)
                m = re.search(r'UNKNOWN_FIELD_NAME.*Unknown field name:\s*\\"([^\\"]+)\\"', msg)
                if not m:
                    m = re.search(r'UNKNOWN_FIELD_NAME.*Unknown field name:\s*"([^"]+)"', msg)
                if requested_fields and m:
                    bad = m.group(1)
                    warn(f"Airtable: champ inconnu '{bad}' -> retrait du filtre fields[] et retry")
                    requested_fields = [f for f in requested_fields if f != bad]
                    continue
                raise

    def _strip_forbidden_fields(self, recs: List[dict]) -> None:
        for rec in recs:
            fields = rec.get("fields")
            if isinstance(fields, dict):
                for f in DO_NOT_WRITE_FIELDS:
                    fields.pop(f, None)

    def create_records(self, records: List[dict]) -> None:
        self._strip_forbidden_fields(records)
        for batch in chunked(records, AIRTABLE_BATCH_SIZE):
            payload = {"records": batch, "typecast": True}
            self._request("POST", self.table_url, data=json.dumps(payload))

    def update_records(self, records: List[dict]) -> None:
        self._strip_forbidden_fields(records)
        for batch in chunked(records, AIRTABLE_BATCH_SIZE):
            payload = {"records": batch, "typecast": True}
            self._request("PATCH", self.table_url, data=json.dumps(payload))

    def delete_records(self, record_ids: List[str]) -> None:
        for batch in chunked(record_ids, AIRTABLE_BATCH_SIZE):
            params = [("records[]", rid) for rid in batch]
            self._request("DELETE", self.table_url, params=params)

# ============================================================
# MAIN
# ============================================================

def main():
    api_token = os.getenv("AIRTABLE_API_TOKEN", "").strip()
    base_id = os.getenv("AIRTABLE_BASE_ID", "").strip()
    table_name = os.getenv("AIRTABLE_CIS_TABLE_NAME", "").strip()

    force_refresh = os.getenv("FORCE_REFRESH", "0").strip() == "1"
    max_cis = os.getenv("MAX_CIS_TO_PROCESS", "").strip()
    max_cis = int(max_cis) if max_cis.isdigit() else 0

    if not api_token or not base_id or not table_name:
        die("Variables manquantes: AIRTABLE_API_TOKEN / AIRTABLE_BASE_ID / AIRTABLE_CIS_TABLE_NAME")

    RUN_ID = run_id_paris()
    MOMENT_REPORT_CSV = os.path.join(REPORT_DIR, f"moment_prise_keywords_{RUN_ID}.csv")
    SUMMARY_JSON = os.path.join(REPORT_DIR, f"run_summary_{RUN_ID}.json")
    init_moment_report_csv(MOMENT_REPORT_CSV)

    atc_labels = load_atc_equivalence_excel(ATC_EQUIVALENCE_FILE)

    info("Téléchargement BDPM CIS ...")
    cis_txt = download_text(BDPM_CIS_URL, encoding="latin-1")
    ok(f"BDPM CIS OK ({len(cis_txt)} chars)")

    info("Téléchargement BDPM CIS_CIP ...")
    cis_cip_txt = download_text(BDPM_CIS_CIP_URL, encoding="latin-1")
    ok(f"BDPM CIS_CIP OK ({len(cis_cip_txt)} chars)")

    info("Téléchargement BDPM COMPO ...")
    compo_txt = download_text(BDPM_COMPO_URL, encoding="latin-1")
    ok(f"BDPM COMPO OK ({len(compo_txt)} chars)")
    compo_map = parse_bdpm_compositions(compo_txt)

    info("Téléchargement BDPM MITM (ATC) ...")
    mitm_txt = download_text(BDPM_MITM_URL, encoding="latin-1")
    ok(f"BDPM MITM OK ({len(mitm_txt)} chars)")
    cis_to_atc = parse_mitm_cis_to_atc(mitm_txt)

    info("Téléchargement BDPM Informations importantes (génération en direct) ...")
    info_imp_txt = download_text(BDPM_INFO_IMPORTANTES_URL, encoding="latin-1")
    ok(f"BDPM Infos importantes OK ({len(info_imp_txt)} chars)")
    cis_to_info_url = parse_info_importantes_cis_to_url(info_imp_txt)

    # ANSM rétrocession (robuste: si KO on continue sans rétrocession)
    ansm_retro_cis: Set[str] = set()
    ansm_link = ""
    try:
        info("Recherche lien Excel ANSM (rétrocession) ...")
        ansm_link = find_ansm_retro_excel_link()
        ok(f"Lien ANSM trouvé: {ansm_link}")

        info("Téléchargement Excel ANSM ...")
        ansm_bytes = download_bytes(ansm_link, timeout_s=140.0)
        ok(f"ANSM Excel OK ({len(ansm_bytes)} bytes)")

        ansm_retro_cis = parse_ansm_retrocession_cis(ansm_bytes, url_hint=ansm_link)
        ok(f"ANSM rétrocession: {len(ansm_retro_cis)} CIS")
    except Exception as e:
        warn(f"ANSM rétrocession indisponible (on continue): {e}")

    cis_map = parse_bdpm_cis(cis_txt)
    cip_map = parse_bdpm_cis_cip(cis_cip_txt)

    at = AirtableClient(api_token, base_id, table_name)

    needed_fields = [
        FIELD_CIS,
        FIELD_RCP,
        FIELD_CIP13,
        FIELD_DISPO,
        FIELD_CPD,
        FIELD_LABO,
        FIELD_SPEC,
        FIELD_FORME,
        FIELD_VOIE,
        FIELD_ATC,
        FIELD_ATC4,        # lecture
        FIELD_ATC_LABEL,   # écriture possible
        FIELD_COMPOSITION, # écriture
        FIELD_COMPOSITION_DETAILS, # écriture
        FIELD_LIEN_INFO_IMPORTANTE, # écriture (URL)
        FIELD_MOMENT_PRISE, # lecture/écriture
        FIELD_INDICATIONS_RCP,
        FIELD_POSOLOGIE_RCP,
        FIELD_INTERACTIONS_RCP,
        FIELD_DATE_REVUE,
    ]

    info("Inventaire Airtable ...")
    records = at.list_all_records(fields=needed_fields)
    ok(f"Enregistrements Airtable: {len(records)}")

    airtable_by_cis: Dict[str, dict] = {}
    for rec in records:
        cis = str(rec.get("fields", {}).get(FIELD_CIS, "")).strip()
        cis = re.sub(r"\D", "", cis)
        if len(cis) == 8:
            airtable_by_cis[cis] = rec

    bdpm_cis_set = set(cis_map.keys())
    airtable_cis_set = set(airtable_by_cis.keys())

    to_create = sorted(list(bdpm_cis_set - airtable_cis_set))
    to_delete = sorted(list(airtable_cis_set - bdpm_cis_set))

    info(f"À créer: {len(to_create)} | À supprimer: {len(to_delete)} | Dans les 2: {len(bdpm_cis_set & airtable_cis_set)}")

    review_ts = now_paris_iso_seconds()

    # CREATE
    if to_create:
        info("Création des enregistrements manquants ...")
        new_recs = []
        for cis in to_create:
            row = cis_map[cis]
            cip = cip_map.get(cis)
            labo = normalize_lab_name(row.titulaire)

            fields: Dict[str, object] = {
                FIELD_CIS: cis,
                FIELD_SPEC: safe_text(row.specialite),
                FIELD_FORME: safe_text(row.forme),
                FIELD_VOIE: safe_text(row.voie_admin),
                FIELD_LABO: labo,
                FIELD_RCP: rcp_link_default(cis),
                FIELD_DATE_REVUE: review_ts,
            }
            if cip and cip.cip13:
                fields[FIELD_CIP13] = cip.cip13

            compo = compo_map.get(cis, "").strip()
            if compo:
                fields[FIELD_COMPOSITION_DETAILS] = compo
                compo_no_acc = strip_accents(compo)
                if compo_no_acc and compo_no_acc.lower() not in compo.lower():
                    fields[FIELD_COMPOSITION] = f"{compo}\n{compo_no_acc}"
                else:
                    fields[FIELD_COMPOSITION] = compo

            atc_code = cis_to_atc.get(cis, "").strip()
            if atc_code:
                fields[FIELD_ATC] = atc_code
                atc4 = atc_level4_from_any(atc_code)
                if atc4:
                    label = atc_labels.get(atc4, "")
                    if label:
                        fields[FIELD_ATC_LABEL] = label

            iu = cis_to_info_url.get(cis, "").strip()
            if iu:
                fields[FIELD_LIEN_INFO_IMPORTANTE] = iu

            fields.pop(FIELD_ATC4, None)
            new_recs.append({"fields": fields})

        at.create_records(new_recs)
        ok(f"Créés: {len(new_recs)}")

        records = at.list_all_records(fields=needed_fields)
        airtable_by_cis = {}
        for rec in records:
            cis = str(rec.get("fields", {}).get(FIELD_CIS, "")).strip()
            cis = re.sub(r"\D", "", cis)
            if len(cis) == 8:
                airtable_by_cis[cis] = rec

    # DELETE
    deleted_count = 0
    if to_delete:
        info("Suppression des enregistrements Airtable absents de BDPM ...")
        ids = [airtable_by_cis[c]["id"] for c in to_delete if c in airtable_by_cis]
        if ids:
            at.delete_records(ids)
            deleted_count += len(ids)
            ok(f"Supprimés: {len(ids)}")

        records = at.list_all_records(fields=needed_fields)
        airtable_by_cis = {}
        for rec in records:
            cis = str(rec.get("fields", {}).get(FIELD_CIS, "")).strip()
            cis = re.sub(r"\D", "", cis)
            if len(cis) == 8:
                airtable_by_cis[cis] = rec

    # ENRICH
    all_cis = sorted(list(set(cis_map.keys()) & set(airtable_by_cis.keys())))
    if max_cis > 0:
        all_cis = all_cis[:max_cis]
        warn(f"MAX_CIS_TO_PROCESS={max_cis} -> {len(all_cis)} CIS traités")

    info("Enrichissement: fiche-info (CPD/dispo) + ATC via MITM + composition + lien info importante + moment de prise (keywords) ...")
    info(f"Revue du jour (timestamp): {review_ts}")

    updates: List[dict] = []
    failures = 0
    start = time.time()

    fiche_checks = 0
    info_added = 0
    atc_added = 0
    moment_checks = 0
    moment_added = 0
    moment_no_kw = 0

    for idx, cis in enumerate(all_cis, start=1):
        if HEARTBEAT_EVERY > 0 and idx % HEARTBEAT_EVERY == 0:
            info(
                f"Heartbeat: {idx}/{len(all_cis)} (CIS={cis}) | fiche checks={fiche_checks} | "
                f"moment checks={moment_checks} | moment added={moment_added} | moment no_kw={moment_no_kw} | "
                f"ATC added={atc_added} | info importante added={info_added} | failures={failures}"
            )

        rec = airtable_by_cis.get(cis)
        if not rec:
            continue

        fields_cur = rec.get("fields", {}) or {}
        upd_fields: Dict[str, object] = {}

        upd_fields[FIELD_DATE_REVUE] = review_ts

        row = cis_map.get(cis)
        if row:
            labo = normalize_lab_name(row.titulaire)
            if labo and str(fields_cur.get(FIELD_LABO, "")).strip() != labo:
                upd_fields[FIELD_LABO] = labo
            if safe_text(row.specialite) and str(fields_cur.get(FIELD_SPEC, "")).strip() != safe_text(row.specialite):
                upd_fields[FIELD_SPEC] = safe_text(row.specialite)
            if safe_text(row.forme) and str(fields_cur.get(FIELD_FORME, "")).strip() != safe_text(row.forme):
                upd_fields[FIELD_FORME] = safe_text(row.forme)
            if safe_text(row.voie_admin) and str(fields_cur.get(FIELD_VOIE, "")).strip() != safe_text(row.voie_admin):
                upd_fields[FIELD_VOIE] = safe_text(row.voie_admin)

        cip = cip_map.get(cis)
        if cip:
            if cip.cip13 and str(fields_cur.get(FIELD_CIP13, "")).strip() != cip.cip13:
                upd_fields[FIELD_CIP13] = cip.cip13

        link_rcp = str(fields_cur.get(FIELD_RCP, "")).strip()
        if not link_rcp:
            link_rcp = rcp_link_default(cis)
            upd_fields[FIELD_RCP] = link_rcp

                # ✅ RCP: extraction keywords (moment de prise) + sections 4.1/4.2/4.5
        cur_moment = str(fields_cur.get(FIELD_MOMENT_PRISE, "")).strip()
        cur_41 = str(fields_cur.get(FIELD_INDICATIONS_RCP, "")).strip()
        cur_42 = str(fields_cur.get(FIELD_POSOLOGIE_RCP, "")).strip()
        cur_45 = str(fields_cur.get(FIELD_INTERACTIONS_RCP, "")).strip()

        need_fetch_rcp = force_refresh or (not cur_moment) or (not cur_41) or (not cur_42) or (not cur_45)

        if need_fetch_rcp and link_rcp:
            rcp_url = set_tab(link_rcp, cis, "rcp")
            try:
                moment_checks += 1
                html_rcp = fetch_html_checked(rcp_url)

                # 1) Moment de prise = mots-clés uniquement
                kw_txt = extract_moment_keywords_from_rcp_html(html_rcp)
                if kw_txt:
                    if kw_txt != cur_moment:
                        upd_fields[FIELD_MOMENT_PRISE] = kw_txt
                        moment_added += 1
                        append_moment_report(MOMENT_REPORT_CSV, RUN_ID, cis, rcp_url, "UPDATED", kw_txt)
                    else:
                        append_moment_report(MOMENT_REPORT_CSV, RUN_ID, cis, rcp_url, "UNCHANGED", kw_txt)
                else:
                    moment_no_kw += 1
                    append_moment_report(MOMENT_REPORT_CSV, RUN_ID, cis, rcp_url, "NO_KEYWORD", "")

                # 2) Sections 4.1 / 4.2 / 4.5
                sections = extract_rcp_sections_from_html(html_rcp)

                sec41 = sections.get("4.1", "").strip()
                sec42 = sections.get("4.2", "").strip()
                sec45 = sections.get("4.5", "").strip()

                if sec41 and sec41 != cur_41:
                    upd_fields[FIELD_INDICATIONS_RCP] = sec41
                if sec42 and sec42 != cur_42:
                    upd_fields[FIELD_POSOLOGIE_RCP] = sec42
                if sec45 and sec45 != cur_45:
                    upd_fields[FIELD_INTERACTIONS_RCP] = sec45

            except PageUnavailable as e:
                warn(f"RCP KO CIS={cis}: {e.detail} ({e.url}) (on continue)")
                append_moment_report(MOMENT_REPORT_CSV, RUN_ID, cis, rcp_url, "RCP_KO", "")
            except Exception as e:
                warn(f"RCP parse KO CIS={cis}: {e} (on continue)")
                append_moment_report(MOMENT_REPORT_CSV, RUN_ID, cis, rcp_url, "RCP_PARSE_KO", "")

        # Composition
        cur_compo = str(fields_cur.get(FIELD_COMPOSITION, "")).strip()
        cur_compo_details = str(fields_cur.get(FIELD_COMPOSITION_DETAILS, "")).strip()
        new_compo = compo_map.get(cis, "").strip()
        if new_compo:
            if new_compo != cur_compo_details:
                upd_fields[FIELD_COMPOSITION_DETAILS] = new_compo

            new_no_acc = strip_accents(new_compo).strip()
            final_compo = new_compo
            if new_no_acc and new_no_acc.lower() not in new_compo.lower():
                final_compo = f"{new_compo}\n{new_no_acc}"
            if final_compo != cur_compo:
                upd_fields[FIELD_COMPOSITION] = final_compo

        # ATC: uniquement via MITM si vide
        cur_atc_raw = str(fields_cur.get(FIELD_ATC, "")).strip()
        if not cur_atc_raw:
            atc_code = cis_to_atc.get(cis, "").strip()
            if atc_code:
                upd_fields[FIELD_ATC] = atc_code
                atc_added += 1
                atc4_tmp = atc_level4_from_any(atc_code)
                if atc4_tmp:
                    label = atc_labels.get(atc4_tmp, "")
                    if label:
                        cur_label = str(fields_cur.get(FIELD_ATC_LABEL, "")).strip()
                        if label != cur_label:
                            upd_fields[FIELD_ATC_LABEL] = label

        # Libellé ATC via ATC4 computed
        cur_atc4 = str(fields_cur.get(FIELD_ATC4, "")).strip()
        cur_label = str(fields_cur.get(FIELD_ATC_LABEL, "")).strip()
        if cur_atc4:
            atc4_norm = atc_level4_from_any(cur_atc4) or cur_atc4.strip().upper()
            label = atc_labels.get(atc4_norm, "")
            if label and label != cur_label:
                upd_fields[FIELD_ATC_LABEL] = label

        # Lien vers information importante
        cur_info_url = str(fields_cur.get(FIELD_LIEN_INFO_IMPORTANTE, "")).strip()
        new_info_url = cis_to_info_url.get(cis, "").strip()
        if new_info_url and new_info_url != cur_info_url:
            upd_fields[FIELD_LIEN_INFO_IMPORTANTE] = new_info_url
            info_added += 1

        # CPD + dispo via fiche-info
        fiche_url = set_tab(link_rcp, cis, "fiche-info")
        cur_cpd = str(fields_cur.get(FIELD_CPD, "")).strip()
        cur_dispo = str(fields_cur.get(FIELD_DISPO, "")).strip()

        need_fetch_fiche = force_refresh or (not cur_cpd) or (not cur_dispo)
        is_retro = cis in ansm_retro_cis

        if need_fetch_fiche:
            try:
                fiche_checks += 1
                cpd_text, is_homeo, reserved_hosp, usage_hosp = analyze_fiche_info(fiche_url)

                if cpd_text and cpd_text != cur_cpd:
                    upd_fields[FIELD_CPD] = cpd_text

                has_taux = cip.has_taux if cip else False
                dispo = compute_disponibilite(
                    has_taux_ville=has_taux,
                    is_ansm_retro=is_retro,
                    is_homeo=is_homeo,
                    reserved_hospital=reserved_hosp,
                    usage_hospital=usage_hosp,
                )
                if dispo != cur_dispo:
                    upd_fields[FIELD_DISPO] = dispo

            except PageUnavailable as e:
                warn(f"Fiche-info KO CIS={cis}: {e.detail} ({e.url}) -> suppression Airtable")

                if updates:
                    at.update_records(updates)
                    ok(f"Batch updates (flush before delete): {len(updates)}")
                    updates = []

                try:
                    at.delete_records([rec["id"]])
                    deleted_count += 1
                    append_deleted_report(cis=cis, reason=f"Page indisponible ({e.status})", url=e.url)
                except Exception as de:
                    failures += 1
                    warn(f"Suppression Airtable impossible CIS={cis}: {de} (on continue)")
                continue

            except Exception as e:
                failures += 1
                warn(f"Enrich KO CIS={cis}: {e} (on continue)")

        upd_fields.pop(FIELD_ATC4, None)
        updates.append({"id": rec["id"], "fields": upd_fields})

        if len(updates) >= UPDATE_FLUSH_THRESHOLD:
            at.update_records(updates)
            ok(f"Batch updates: {len(updates)}")
            updates = []

        if idx % 1000 == 0:
            elapsed = time.time() - start
            rate = idx / elapsed if elapsed > 0 else 0
            remaining = (len(all_cis) - idx) / rate if rate > 0 else 0
            info(
                f"Progress {idx}/{len(all_cis)} | {rate:.2f} CIS/s | échecs={failures} | supprimés={deleted_count} "
                f"| fiche checks={fiche_checks} | moment checks={moment_checks} | moment added={moment_added} | no_kw={moment_no_kw} "
                f"| reste ~{int(remaining)}s"
            )

    if updates:
        at.update_records(updates)
        ok(f"Updates finaux: {len(updates)}")

    # Résumé JSON
    ensure_report_dir()
    summary = {
        "run_id": RUN_ID,
        "timestamp_paris": now_paris_iso_seconds(),
        "total_cis_processed": len(all_cis),
        "moment_checks": moment_checks,
        "moment_added": moment_added,
        "moment_no_keyword": moment_no_kw,
        "fiche_checks": fiche_checks,
        "atc_added": atc_added,
        "info_added": info_added,
        "failures": failures,
        "deleted_count": deleted_count,
        "moment_report_csv": MOMENT_REPORT_CSV,
    }
    with open(SUMMARY_JSON, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    try_git_commit_reports()

    ok(
        f"Terminé. échecs={failures} | supprimés={deleted_count} | "
        f"fiche checks={fiche_checks} | moment checks={moment_checks} | moment added={moment_added} | no_kw={moment_no_kw} | "
        f"ATC added={atc_added} | info importante added={info_added} | "
        f"reports: {MOMENT_REPORT_CSV} ; {SUMMARY_JSON}"
    )

if __name__ == "__main__":
    main()
